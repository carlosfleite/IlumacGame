# -*- coding: utf-8 -*-
"""
Painel do marketing — dados de contato dos participantes, com exportação
em CSV, Excel e PDF. Protegido por senha; ninguém mais no totem acessa.

config/admin.json guarda o hash da senha (nunca a senha em texto puro) e
a chave usada para assinar o cookie de sessão do Flask. NÃO é versionado
(ver .gitignore) — é gerado sozinho no primeiro boot, com uma senha
aleatória registrada no log, pra ninguém precisar configurar nada antes
do evento só pra essa tela funcionar. Pra trocar a senha depois, use
tools/definir_senha_admin.py (ou apague o arquivo pra gerar outra).
"""
import csv
import hashlib
import io
import json
import logging
import os
import secrets
from datetime import datetime
from functools import wraps

from flask import Blueprint, Response, jsonify, redirect, render_template, request, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from fpdf import FPDF

from database import CONFIG_DIR, get_connection, sql_melhor_tentativa

ADMIN_JSON = os.path.join(CONFIG_DIR, "admin.json")

log = logging.getLogger(__name__)

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")


def _carregar_ou_criar_credenciais():
    if os.path.exists(ADMIN_JSON):
        with open(ADMIN_JSON, encoding="utf-8") as fp:
            return json.load(fp)

    senha = secrets.token_urlsafe(9)
    dados = {
        "senha_hash": hashlib.sha256(senha.encode("utf-8")).hexdigest(),
        "secret_key": secrets.token_hex(32),
    }
    os.makedirs(CONFIG_DIR, exist_ok=True)
    with open(ADMIN_JSON, "w", encoding="utf-8") as fp:
        json.dump(dados, fp, indent=2)

    mensagem = (
        "Senha do painel admin gerada automaticamente: %s\n"
        "ANOTE ESTA SENHA AGORA. Ela nao fica salva em texto puro em "
        "nenhum lugar (config/admin.json guarda so o hash) — se perder, "
        "rode tools/definir_senha_admin.py pra definir uma nova."
    ) % senha

    # log.warning sozinho não é confiável aqui: este código roda na
    # IMPORTAÇÃO do módulo, ou seja, antes de run.py configurar o
    # logging (que só acontece dentro de main()). Nesse instante o
    # logger ainda não tem handler de arquivo — a mensagem podia só
    # aparecer (e sumir) na janela preta do console. print() + arquivo
    # garantem que a senha sobrevive mesmo que ninguém esteja olhando
    # pra tela no segundo exato em que o totem liga.
    print("\n" + "=" * 70 + "\n" + mensagem + "\n" + "=" * 70 + "\n")
    log.warning(mensagem)
    try:
        log_dir = os.path.join(os.path.dirname(CONFIG_DIR), "logs")
        os.makedirs(log_dir, exist_ok=True)
        caminho = os.path.join(log_dir, "SENHA_ADMIN_GERADA_UMA_VEZ.txt")
        with open(caminho, "w", encoding="utf-8") as fp:
            fp.write(mensagem + "\nApague este arquivo depois de anotar a senha.\n")
    except OSError:
        pass  # não impede o totem de subir por causa disso

    return dados


_CREDENCIAIS = _carregar_ou_criar_credenciais()
SECRET_KEY = _CREDENCIAIS["secret_key"]


def _senha_correta(senha):
    digest = hashlib.sha256((senha or "").encode("utf-8")).hexdigest()
    return secrets.compare_digest(digest, _CREDENCIAIS["senha_hash"])


def _requer_admin(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("admin"):
            if "/exportar/" in request.path:
                return jsonify({"ok": False, "erro": "Não autenticado."}), 401
            return redirect(url_for("admin.login"))
        return fn(*args, **kwargs)
    return wrapper


@admin_bp.route("/login", methods=["GET", "POST"])
def login():
    erro = None
    if request.method == "POST":
        if _senha_correta(request.form.get("senha")):
            session.clear()
            session["admin"] = True
            session.permanent = True
            return redirect(url_for("admin.painel"))
        erro = "Senha incorreta."
    return render_template("admin_login.html", erro=erro)


@admin_bp.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("admin.login"))


# ---------------------------------------------------------------------------
# Dados
# ---------------------------------------------------------------------------

def _buscar_participantes(conn):
    """
    Um cadastro por linha, com a melhor pontuação e o prêmio ganho (se
    houver). Participantes que se cadastraram mas não terminaram o quiz
    também entram — para o marketing a lista de leads importa inteira,
    não só quem jogou até o fim.
    """
    return conn.execute(
        """
        SELECT
            p.id, p.nome, p.telefone, p.email,
            p.consentimento_lgpd, p.data_cadastro,
            best.pontuacao, best.tempo_total_ms, best.data_hora,
            pr.nome AS premio_nome
        FROM participantes p
        LEFT JOIN (%s) best ON best.participante_id = p.id
        LEFT JOIN quiz_premios pr ON pr.id = best.premio_id
        ORDER BY p.data_cadastro DESC
        """ % sql_melhor_tentativa()
    ).fetchall()


def _tempo_legivel(ms):
    if ms is None:
        return ""
    s = int(ms) // 1000
    return "%02d:%02d" % (s // 60, s % 60)


def _nome_arquivo(extensao):
    return "ilumac_participantes_%s.%s" % (datetime.now().strftime("%Y-%m-%d_%H%M"), extensao)


# ---------------------------------------------------------------------------
# Painel
# ---------------------------------------------------------------------------

@admin_bp.route("/")
@_requer_admin
def painel():
    conn = get_connection()
    try:
        linhas = _buscar_participantes(conn)
    finally:
        conn.close()

    total = len(linhas)
    consentiram = sum(1 for r in linhas if r["consentimento_lgpd"])
    jogaram = sum(1 for r in linhas if r["pontuacao"] is not None)

    participantes = [
        {
            "nome": r["nome"],
            "telefone": r["telefone"] or "",
            "email": r["email"] or "",
            "consentimento": bool(r["consentimento_lgpd"]),
            "data_cadastro": r["data_cadastro"] or "",
            "pontuacao": r["pontuacao"],
            "tempo": _tempo_legivel(r["tempo_total_ms"]),
            "premio": r["premio_nome"] or "",
            "jogou": r["pontuacao"] is not None,
        }
        for r in linhas
    ]

    return render_template(
        "admin_painel.html",
        participantes=participantes,
        total=total,
        consentiram=consentiram,
        jogaram=jogaram,
    )


# ---------------------------------------------------------------------------
# Exportações
# ---------------------------------------------------------------------------

@admin_bp.route("/exportar/participantes.csv")
@_requer_admin
def exportar_csv():
    """
    Delimitador ';' e BOM UTF-8: é o que faz o Excel em português abrir
    o arquivo direto, com acentos corretos, sem passar pelo assistente de
    importação. Vírgula como separador de campo colide com a vírgula
    decimal do Excel PT-BR e tudo cai numa coluna só.
    """
    conn = get_connection()
    try:
        linhas = _buscar_participantes(conn)
    finally:
        conn.close()

    buffer = io.StringIO()
    escritor = csv.writer(buffer, delimiter=";")
    escritor.writerow([
        "nome", "telefone", "email", "consentimento_lgpd", "data_cadastro",
        "pontuacao", "tempo_total", "premio", "jogou",
    ])
    for r in linhas:
        escritor.writerow([
            r["nome"],
            r["telefone"] or "",
            r["email"] or "",
            "sim" if r["consentimento_lgpd"] else "nao",
            r["data_cadastro"] or "",
            r["pontuacao"] if r["pontuacao"] is not None else "",
            _tempo_legivel(r["tempo_total_ms"]),
            r["premio_nome"] or "",
            "sim" if r["pontuacao"] is not None else "nao",
        ])

    return Response(
        buffer.getvalue().encode("utf-8-sig"),
        mimetype="text/csv",
        headers={"Content-Disposition": 'attachment; filename="%s"' % _nome_arquivo("csv")},
    )


@admin_bp.route("/exportar/participantes.xlsx")
@_requer_admin
def exportar_xlsx():
    conn = get_connection()
    try:
        linhas = _buscar_participantes(conn)
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Participantes"

    cabecalho = [
        "Nome", "Telefone", "E-mail", "Consentimento LGPD",
        "Data de cadastro", "Pontuação", "Tempo", "Prêmio", "Jogou",
    ]
    ws.append(cabecalho)
    for celula in ws[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="C31922")

    for r in linhas:
        ws.append([
            r["nome"],
            r["telefone"] or "",
            r["email"] or "",
            "Sim" if r["consentimento_lgpd"] else "Não",
            r["data_cadastro"] or "",
            r["pontuacao"] if r["pontuacao"] is not None else None,
            _tempo_legivel(r["tempo_total_ms"]),
            r["premio_nome"] or "",
            "Sim" if r["pontuacao"] is not None else "Não",
        ])

    for coluna in ws.columns:
        valores = [str(c.value) for c in coluna if c.value is not None]
        largura = max((len(v) for v in valores), default=10) + 2
        ws.column_dimensions[coluna[0].column_letter].width = min(largura, 42)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="%s"' % _nome_arquivo("xlsx")},
    )


def _pdf_seguro(texto, tamanho_max=None):
    """
    A fonte core do fpdf2 (Helvetica) só lida com Latin-1 — cobre os
    acentos do português, mas não travessão tipográfico, emoji etc.
    Substitui o que não encaixa por '?' em vez de estourar a exportação.
    """
    texto = str(texto or "")
    texto = texto.encode("latin-1", "replace").decode("latin-1")
    if tamanho_max and len(texto) > tamanho_max:
        texto = texto[: tamanho_max - 1] + "…".encode("latin-1", "replace").decode("latin-1")
    return texto


_COLUNAS_PDF = [
    ("Nome", 42),
    ("Telefone", 28),
    ("E-mail", 52),
    ("LGPD", 14),
    ("Cadastro", 28),
    ("Pontos", 14),
    ("Prêmio", 40),
    ("Jogou", 14),
]


@admin_bp.route("/exportar/participantes.pdf")
@_requer_admin
def exportar_pdf():
    conn = get_connection()
    try:
        linhas = _buscar_participantes(conn)
    finally:
        conn.close()

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, _pdf_seguro("Participantes - Quiz SDAI Ilumac"), ln=1)
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, _pdf_seguro("Gerado em %s" % datetime.now().strftime("%d/%m/%Y %H:%M")), ln=1)
    pdf.ln(2)

    def cabecalho_tabela():
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(195, 25, 34)
        pdf.set_text_color(255, 255, 255)
        for titulo, largura in _COLUNAS_PDF:
            pdf.cell(largura, 8, _pdf_seguro(titulo), border=1, fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "", 8)

    cabecalho_tabela()
    for i, r in enumerate(linhas):
        if pdf.get_y() > 190:  # perto do rodapé: nova página com cabeçalho de novo
            pdf.add_page()
            cabecalho_tabela()
        valores = [
            _pdf_seguro(r["nome"], 34),
            _pdf_seguro(r["telefone"], 20),
            _pdf_seguro(r["email"], 42),
            "Sim" if r["consentimento_lgpd"] else "Não",
            _pdf_seguro((r["data_cadastro"] or "")[:16], 22),
            str(r["pontuacao"]) if r["pontuacao"] is not None else "",
            _pdf_seguro(r["premio_nome"], 32),
            "Sim" if r["pontuacao"] is not None else "Não",
        ]
        pdf.set_fill_color(246, 246, 246)
        preencher = i % 2 == 1
        for (titulo, largura), valor in zip(_COLUNAS_PDF, valores):
            pdf.cell(largura, 7, valor, border=1, fill=preencher)
        pdf.ln()

    return Response(
        bytes(pdf.output()),
        mimetype="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="%s"' % _nome_arquivo("pdf")},
    )
